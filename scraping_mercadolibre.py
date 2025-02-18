from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

# Configurar el navegador
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)

# URL de la página de inicio
# url = "https://listado.mercadolibre.com.ve/inmuebles/casas/venta/lara/_PriceRange_0USD-27000USD"
def find_next_url(filename="master_state.xlsx", sheet_name="Sheet1"):
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]
    # Encontrar URL a procesar
    # Encontrar la primera fila vacía
    row = 1
    while sheet.cell(row=row, column=2).value is not None:
        row += 1
    url_to_process = sheet.cell(row=row, column=3).value
    sheet.cell(row=row, column=2, value=1)
    workbook.save(filename)
    return url_to_process

# Función para obtener y contar los enlaces de las casas en la página actual
def get_house_links():
    house_links = []
    houses = driver.find_elements(By.XPATH, '//li[contains(@class, "ui-search-layout__item")]//a')
    for house in houses:
        house_links.append(house.get_attribute("href"))
    
    return house_links

# Función para guardar los enlaces en un archivo de Excel
def save_links_to_excel(links, filename="master_state.xlsx", sheet_name="Details"):
    workbook = load_workbook(filename)
    sheet = workbook[sheet_name]

    # Encontrar la primera fila vacía
    row = 1
    while sheet.cell(row=row, column=2).value is not None:
        row += 1

    # Escribir los enlaces en la hoja
    for link in links:
        sheet.cell(row=row, column=2, value=link)
        row += 1

    # Guardar el archivo
    workbook.save(filename)
url = find_next_url()
# Iniciar navegación
driver.get(url)
time.sleep(10)  # Esperar un poco para cargar la página

# Obtener y contar los enlaces de las casas de la página actual
current_house_links = get_house_links()
print(f"Total de casas en la página actual: {len(current_house_links)}")

# Guardar los enlaces en el archivo de Excel
save_links_to_excel(current_house_links)

# Buscar el enlace del botón "Siguiente" y obtener su URL
next_button = driver.find_element(By.XPATH, "//li[contains(@class, 'andes-pagination__button--next')]/a")
next_url = next_button.get_attribute("href")

while next_url:
    # Abrir la URL del botón "Siguiente"
    driver.get(next_url)
    time.sleep(7)  # Esperar un poco para cargar la página

    # Obtener y contar los enlaces de las casas de la página actual
    current_house_links = get_house_links()
    print(f"Total de casas en la página actual: {len(current_house_links)}")

    # Guardar los enlaces en el archivo de Excel
    save_links_to_excel(current_house_links)

    # Buscar el enlace del siguiente botón "Siguiente" y obtener su URL
    try:
        next_button = driver.find_element(By.XPATH, "//li[contains(@class, 'andes-pagination__button--next')]/a")
        next_url = next_button.get_attribute("href")
    except:
        print("INFO: No se encontró el botón 'Siguiente'.")
        next_url = None

# Cerrar el navegador
driver.quit()
