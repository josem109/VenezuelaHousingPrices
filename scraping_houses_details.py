import openpyxl
import random
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import re
from xpath_constants import (
    XPATH_DATA,
    XPATH_LOCATION,XPATH_METERS,XPATH_PRICE,
    XPATH_METERS2,XPATH_PRICE2
)

# Solicitar el número de líneas a procesar
num_lines_to_process = int(input("Ingrese el número de líneas a procesar: "))
# Start time measurement
start_time = time.time()

# Abrir el archivo de Excel
workbook = openpyxl.load_workbook("master_state.xlsx")
sheet = workbook["Details"]
#data_sheet = workbook["Data"]
data_sheet = workbook["DataReg"]
# Configurar el navegador
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)

def remove_trailing_comma(input_str):
  # Check if the string ends with a comma
  if input_str[-1] == ',':
    # Remove the trailing comma
    return input_str[:-1]
  else:
    # Return the original string if it doesn't end with a comma
    return input_str
# Función para extraer la información usando Selenium
def extract_data(url):
    driver.get(url)
    time.sleep(random.uniform(1, 2))  # Esperar a que la página cargue completamente
    driver.refresh()  # Refrescar la página
    time.sleep(random.uniform(1, 2))
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    try:
        show_all_button = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "span[data-testid='action-collapsable-target'][role='button']"))
        )
        show_all_button.click()
    except Exception as e:
        print(f"No se pudo hacer clic en 'Ver todas las características': {e}")

    # Intentar extraer todos los campos, manejar excepciones para campos no encontrados
    def safe_find_element(xpath):
        try:
            element = driver.find_element(By.XPATH, xpath)
            return element.text if isinstance(element.text, str) else element.text[0]
        except:
            return "N/A"
        # Extraer valores adicionales
    location = safe_find_element(XPATH_LOCATION)
    meters = safe_find_element(XPATH_METERS)
    price = safe_find_element(XPATH_PRICE)
    if meters == 'N/A':
        meters = safe_find_element(XPATH_METERS2)
    if price == 'N/A':
        price = safe_find_element(XPATH_PRICE2)
    property_info = safe_find_element(XPATH_DATA)
    #print(property_info)
    # Define excluded values
    # Define the value to exclude
    excluded_values = ["Principales", "Servicios", "Comodidades y equipamiento","Condiciones especiales"]  # Define excluded values
    
    #attributes = []
    attributes = [("Location", location), ("Meters", meters), ("Price", price), ("URL", url)]
    is_attribute_name = True
    
    lines = property_info.splitlines()
    #print(f"{lines[0]}: Primer valor")
    for index, line in enumerate(lines):
        is_excluded = False
        #print(f"Index: {index}, Line: {line}")
        #if line and not any(excluded_value in line for excluded_value in excluded_values):
        if line and not any(excluded_value == line for excluded_value in excluded_values):
            # Check if any excluded value is a substring of the line
            if "Seguridad".lower() == line.lower():
            # Check if next line exists (index + 1 is within list bounds)
                if index + 1 < len(lines):
                    next_line = lines[index + 1]

                # Check if next line is either "Sí" or "No" (case insensitive)
                if next_line.lower() in ["sí", "no"]:
                    #print(f"Current line: {line}, Next line: {next_line}")
                    is_excluded = False
                else:
                    is_excluded = True
                    is_attribute_name = True
             # Check if current line is "Ambientes" (case insensitive)
            elif "Ambientes".lower() == line.lower():
                # Check if next line exists (index + 1 is within list bounds)
                if index + 1 < len(lines):
                    next_line = lines[index + 1]

                    # Check if next line is non-numeric
                    try:
                        float(next_line)
                        is_excluded = False  # Set excluded even if numeric
                        
                    except ValueError:
                        is_excluded = True
                        #print(f"Current line: {line}, Next line: {next_line}")
                        is_attribute_name = True
                else:
                    is_excluded = True
            if is_excluded == False:
                if is_attribute_name:
                    attribute_name = line.strip()
                    is_attribute_name = False
                else:
                    attribute_value = line.strip()
                    attributes.append((attribute_name, attribute_value))
                    is_attribute_name = True

    for attribute_name, attribute_value in attributes:
        print(f"{attribute_name}: {attribute_value}")
    data = {}
    
    attribute_names = [attribute_name for attribute_name, _ in attributes]
    attribute_values = [attribute_value for _, attribute_value in attributes]

    existing_columns = {data_sheet.cell(row=1, column=col).value: col for col in range(1, data_sheet.max_column + 1)}

    # Agregar nuevas columnas si es necesario
    for attribute_name in attribute_names:
        if attribute_name not in existing_columns:
            new_col = data_sheet.max_column + 1
            data_sheet.cell(row=1, column=new_col).value = attribute_name
            existing_columns[attribute_name] = new_col
            
    # Encontrar la primera fila vacía para insertar valores
    first_empty_row = data_sheet.max_row + 1
    for row in range(2, data_sheet.max_row + 1):
        if all(data_sheet.cell(row=row, column=col).value is None for col in range(1, data_sheet.max_column + 1)):
            first_empty_row = row
            break

    # Insertar valores en las columnas correspondientes
    for attribute_name, attribute_value in zip(attribute_names, attribute_values):
        col = existing_columns[attribute_name]
        data_sheet.cell(row=first_empty_row, column=col).value = attribute_value
    #End Testing
# Encontrar la primera fila vacía en la columna A y procesar desde ahí
processed_count = 0
start_row = 2
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value is None:
        start_row = row
        break

# Procesar las URLs a partir de la primera fila vacía encontrada
for row in range(start_row, sheet.max_row + 1):
    if processed_count >= num_lines_to_process:
        break
    cell_value = sheet.cell(row=row, column=2).value  # Columna B
    if cell_value:
        url = cell_value
        extract_data(url)
        sheet.cell(row=row, column=1).value = 1  # Marcar la fila como procesada en la columna A
        processed_count += 1
        # Calcular y mostrar el progreso
        progress = (processed_count / num_lines_to_process) * 100
        print(f"Progreso: {progress:.2f}%")


# End time measurement
end_time = time.time()
total_time = end_time - start_time

# Calculate execution time in minutes and seconds
minutes, seconds = divmod(total_time, 60)

# Print final message with processing time
print(f"Se procesaron {processed_count} registros en {minutes:.0f} minutos y {seconds:.2f} segundos.")
# Guardar el archivo de Excel
workbook.save("master_state.xlsx")

# Cerrar el navegador
driver.quit()
